"""
优化版FastAPI路由 - 使用预计算表实现超快速查询
性能提升: 10-50倍

AI分析: 使用DeepSeek-R1 + 多级降级策略
"""
from fastapi import APIRouter, HTTPException, Query
from typing import List, Dict, Any, Optional
from backend.analytics.target_buyer_analyzer import TargetBuyerAnalyzer
from backend.ai.analyzer_orchestrator import get_analyzer_orchestrator
from backend.database import BuyerQueries

router = APIRouter(prefix="/api/v2", tags=["target_buyers"])
analyzer = TargetBuyerAnalyzer()
ai_orchestrator = get_analyzer_orchestrator()


@router.get("/")
async def root():
    """API health check"""
    return {
        "status": "ok",
        "service": "SmokeSignal Analytics API v2 (Optimized)",
        "version": "2.0.0",
        "performance": "Using precomputed table - 10-50x faster"
    }


@router.get("/buyers")
async def get_all_buyers(
    search: Optional[str] = Query(None, description="昵称模糊搜索"),
    buyer_type: Optional[List[str]] = Query(None, description="买家类型: SMOKER/VIC/BOTH"),
    vip_level: Optional[List[str]] = Query(None, description="VIP等级: V3/V2/V1/V0/Non-VIP"),
    channel: Optional[List[str]] = Query(None, description="渠道: DTC/PFS"),
    last_purchase_after: Optional[str] = Query(None, description="最后购买日期筛选 (YYYY-MM-DD)"),
    chat_status: Optional[str] = Query(None, description="聊天状态: chatted/no_chat"),
    sort_by: str = Query('last_purchase', description="排序字段: last_purchase/l6m_netsales/vip_level"),
    limit: int = Query(100, ge=1, le=1000, description="返回数量"),
    offset: int = Query(0, ge=0, description="偏移量"),
    include_total: bool = Query(False, description="是否返回总数(可能较慢)")
) -> Dict[str, Any]:
    """
    获取所有目标买家列表(超快!)

    性能: < 0.5秒 (优化前: 10-30秒)

    支持的筛选条件:
    - search: 昵称模糊搜索
    - buyer_type: 买家类型 (SMOKER/VIC/BOTH)
    - vip_level: VIP等级 (V3/V2/V1/V0/Non-VIP)
    - channel: 渠道 (DTC/PFS)
    - last_purchase_after: 最后购买日期筛选
    - chat_status: 聊天状态 (chatted/no_chat)
    - sort_by: 排序字段 (last_purchase/l6m_netsales/vip_level)
    """
    try:
        result = analyzer.get_all_buyers(
            search=search,
            buyer_type=buyer_type,
            vip_level=vip_level,
            channel=channel,
            last_purchase_after=last_purchase_after,
            chat_status=chat_status,
            sort_by=sort_by,
            limit=limit,
            offset=offset,
            include_total=include_total
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/count")
async def get_buyers_count(
    buyer_type: Optional[List[str]] = Query(None, description="买家类型: SMOKER/VIC/BOTH"),
    vip_level: Optional[List[str]] = Query(None, description="VIP等级: V3/V2/V1/V0/Non-VIP"),
    channel: Optional[List[str]] = Query(None, description="渠道: DTC/PFS"),
    last_purchase_after: Optional[str] = Query(None, description="最后购买日期筛选 (YYYY-MM-DD)"),
    chat_status: Optional[str] = Query(None, description="聊天状态: chatted/no_chat")
) -> Dict[str, Any]:
    try:
        total = analyzer.get_buyers_count(
            buyer_type=buyer_type,
            vip_level=vip_level,
            channel=channel,
            last_purchase_after=last_purchase_after,
            chat_status=chat_status
        )
        return {"total": total}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/{user_nick}")
async def get_buyer_profile(
    user_nick: str,
    include_ai: bool = Query(False, description="是否包含AI分析")
) -> Dict[str, Any]:
    """
    获取买家360°详情(超快!)

    性能: < 0.1秒 (优化前: 2-5秒)

    Args:
        user_nick: 买家昵称
        include_ai: 是否包含AI分析(会慢一些)
    """
    try:
        import logging
        # 从预计算表获取基本信息(超快)
        profile = analyzer.get_buyer_profile(user_nick)
        logging.info(f"[DEBUG] get_buyer_profile: user_nick={user_nick}, include_ai={include_ai}")

        if not profile:
            raise HTTPException(status_code=404, detail=f"买家 {user_nick} 不存在")

        # 如果需要AI分析,添加AI生成的洞察
        if include_ai:
            logging.info(f"[DEBUG] get_buyer_profile: calling _add_ai_analysis for {user_nick}")
            profile = await _add_ai_analysis(profile)

        return profile

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/dashboard/metrics")
async def get_dashboard_metrics() -> Dict[str, Any]:
    """
    获取Dashboard汇总指标(超快!)

    性能: < 0.1秒 (优化前: 5-15秒)
    """
    try:
        metrics = analyzer.get_dashboard_metrics()
        return metrics
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/type/{buyer_type}")
async def get_buyers_by_type(
    buyer_type: str,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    按买家类型筛选

    性能: < 0.1秒 (使用索引)

    Args:
        buyer_type: 买家类型 (SMOKER/VIC/BOTH)
    """
    try:
        if buyer_type not in ['SMOKER', 'VIC', 'BOTH']:
            raise HTTPException(
                status_code=400,
                detail=f"无效的buyer_type: {buyer_type}. 必须是 SMOKER, VIC 或 BOTH"
            )

        buyers = analyzer.get_buyers_by_type(buyer_type, limit, offset)
        return buyers
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/vic")
async def get_vic_buyers(
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    获取VIC买家列表(按VIP等级排序)

    性能: < 0.1秒 (使用vip_level索引)
    """
    try:
        buyers = analyzer.get_vic_buyers(limit, offset)
        return buyers
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/smoker")
async def get_smoker_buyers(
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    获取Smoker买家列表(按消费金额排序)

    性能: < 0.1秒 (使用l6m_netsales索引)
    """
    try:
        buyers = analyzer.get_smoker_buyers(limit, offset)
        return buyers
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/churn-risk/{risk_level}")
async def get_churn_risk_buyers(
    risk_level: str,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    获取流失风险买家

    性能: < 0.1秒 (使用churn_risk索引)

    Args:
        risk_level: 风险等级 (高/中/低)
    """
    try:
        if risk_level not in ['高', '中', '低']:
            raise HTTPException(
                status_code=400,
                detail=f"无效的risk_level: {risk_level}. 必须是 高, 中 或 低"
            )

        buyers = analyzer.get_churn_risk_buyers(risk_level, limit, offset)
        return buyers
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/high-value")
async def get_high_value_buyers(
    min_l6m_netsales: float = Query(5000, ge=0, description="最小近6个月消费金额"),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    获取高价值买家(L6M消费 >= 阈值)

    性能: < 0.1秒 (使用l6m_netsales索引)
    """
    try:
        buyers = analyzer.get_high_value_buyers(min_l6m_netsales, limit, offset)
        return buyers
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/vip-level/{vip_level}")
async def get_buyers_by_vip_level(
    vip_level: str,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    按VIP等级筛选买家

    性能: < 0.1秒 (使用vip_level索引)

    Args:
        vip_level: VIP等级 (V3/V2/V1/V0/Non-VIP)
    """
    try:
        if vip_level not in ['V3', 'V2', 'V1', 'V0', 'Non-VIP']:
            raise HTTPException(
                status_code=400,
                detail=f"无效的vip_level: {vip_level}. 必须是 V3, V2, V1, V0 或 Non-VIP"
            )

        buyers = analyzer.get_buyers_by_vip_level(vip_level, limit, offset)
        return buyers
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats/channel")
async def get_channel_stats() -> List[Dict[str, Any]]:
    """
    按渠道统计(DTC/PFS)

    性能: < 0.1秒 (使用channel索引)
    """
    try:
        stats = analyzer.get_channel_stats()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ai/status")
async def get_ai_system_status() -> Dict[str, Any]:
    """
    获取AI分析系统状态

    返回:
    - 可用的AI模型
    - 成本统计
    - 缓存状态
    """
    try:
        from backend.monitoring.cost_monitor import get_cost_monitor
        from backend.utils.cache_manager import get_cache_manager
        from backend.config import settings

        cost_monitor = get_cost_monitor()
        cache_manager = get_cache_manager()

        # 获取成本统计
        cost_summary = cost_monitor.get_daily_summary()

        # 获取缓存统计
        cache_stats = cache_manager.get_stats()

        # 检查模型可用性
        models = {
            "deepseek_available": bool(settings.deepseek_api_key),
            "zhipu_available": bool(settings.zhipu_api_key)
        }

        return {
            "status": "ok",
            "models": models,
            "cost": cost_summary,
            "cache": cache_stats,
            "config": {
                "cache_ttl_days": settings.ai_cache_ttl_days,
                "cache_enabled": settings.ai_enable_cache
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ai/cost-stats/daily")
async def get_cost_stats_daily() -> Dict[str, Any]:
    """
    获取今日成本统计 + 小时成本明细

    返回:
    - 每日汇总（调用次数、总成本、预算使用率）
    - 按模型统计
    - 小时成本明细（用于图表）
    """
    try:
        from backend.monitoring.cost_monitor import get_cost_monitor
        from datetime import date

        cost_monitor = get_cost_monitor()

        # 每日汇总
        daily_summary = cost_monitor.get_daily_summary()

        # 小时成本明细
        hourly_costs = cost_monitor.get_hourly_costs_today()

        return {
            "date": str(date.today()),
            "daily": daily_summary,
            "hourly": hourly_costs
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ai/cost-stats/monthly")
async def get_cost_stats_monthly(
    year: int = Query(None, description="年份（默认今年）"),
    month: int = Query(None, description="月份（默认本月）")
) -> Dict[str, Any]:
    """
    获取月度成本统计

    Args:
        year: 年份（默认今年）
        month: 月份（默认本月）
    """
    try:
        from backend.monitoring.cost_monitor import get_cost_monitor
        from datetime import date

        cost_monitor = get_cost_monitor()

        # 默认为今年本月
        today = date.today()
        year = year or today.year
        month = month or today.month

        # 月度汇总
        monthly_summary = cost_monitor.get_monthly_summary(year, month)

        return monthly_summary

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ai/cost-stats/by-model")
async def get_cost_stats_by_model(
    start_date: str = Query(..., description="开始日期 (YYYY-MM-DD)"),
    end_date: str = Query(..., description="结束日期 (YYYY-MM-DD)")
) -> Dict[str, Any]:
    """
    获取按模型分组的成本统计

    Args:
        start_date: 开始日期 (YYYY-MM-DD)
        end_date: 结束日期 (YYYY-MM-DD)
    """
    try:
        from backend.monitoring.cost_monitor import get_cost_monitor
        from datetime import datetime

        cost_monitor = get_cost_monitor()

        # 解析日期
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()

        # 按模型统计
        model_stats = cost_monitor.get_cost_by_model(start, end)

        return {
            "period": {
                "start": str(start),
                "end": str(end)
            },
            "models": model_stats
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"日期格式错误: {e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/actionable-customers")
async def get_actionable_customers(
    limit: int = Query(50, ge=1, le=500, description="每个类别返回数量")
) -> Dict[str, List[Dict[str, Any]]]:
    """
    获取需要优先处理的客户(可操作客户列表)

    整合流失风险、高价值等需要关注的客户

    性能: < 0.5秒 (多次索引查询)
    """
    try:
        customers = analyzer.get_actionable_customers(limit)
        return customers
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/priority-customers")
async def get_priority_customers(
    channel: Optional[List[str]] = Query(None, description="渠道筛选: DTC/PFS"),
    buyer_type: Optional[List[str]] = Query(None, description="买家类型: SMOKER/VIC/BOTH"),
    follow_priority: Optional[List[str]] = Query(None, description="跟进优先级: 紧急/高/中/低"),
    sentiment_label: Optional[List[str]] = Query(None, description="情感标签: Positive/Neutral/Negative"),
    has_chat: Optional[str] = Query(None, description="聊天状态: yes/no"),
    use_default_filter: bool = Query(True, description="使用默认筛选(紧急/高优先级 OR 负面情感)"),
    limit: int = Query(20, ge=1, le=100, description="每页数量"),
    offset: int = Query(0, ge=0, description="偏移量"),
    include_total: bool = Query(True, description="是否返回总数")
) -> Dict[str, Any]:
    """
    获取优先关注客户列表(带AI画像)

    默认筛选: follow_priority IN ('紧急', '高') OR sentiment_label = 'Negative'

    支持的筛选条件:
    - channel: 渠道 (DTC/PFS)
    - buyer_type: 买家类型 (SMOKER/VIC/BOTH)
    - follow_priority: 跟进优先级 (紧急/高/中/低)
    - sentiment_label: 情感标签 (Positive/Neutral/Negative)
    - has_chat: 聊天状态 (yes/no)

    返回字段包含AI画像:
    - persona_key_interests: 关键兴趣点
    - persona_pain_points: 痛点
    - persona_recommended_action: 推荐行动

    性能: < 1秒 (JOIN + 索引查询)
    """
    try:
        # 获取客户列表
        customers = analyzer.get_priority_customers(
            channel=channel,
            buyer_type=buyer_type,
            follow_priority=follow_priority,
            sentiment_label=sentiment_label,
            has_chat=has_chat,
            use_default_filter=use_default_filter,
            limit=limit,
            offset=offset
        )

        # 获取总数
        total = None
        if include_total:
            total = analyzer.get_priority_customers_count(
                channel=channel,
                buyer_type=buyer_type,
                follow_priority=follow_priority,
                sentiment_label=sentiment_label,
                has_chat=has_chat,
                use_default_filter=use_default_filter
            )

        return {
            "customers": customers,
            "total": total,
            "limit": limit,
            "offset": offset
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/priority-customers/export")
async def export_priority_customers_csv(
    channel: Optional[List[str]] = Query(None, description="渠道筛选: DTC/PFS"),
    buyer_type: Optional[List[str]] = Query(None, description="买家类型: SMOKER/VIC/BOTH"),
    follow_priority: Optional[List[str]] = Query(None, description="跟进优先级: 紧急/高/中/低"),
    sentiment_label: Optional[List[str]] = Query(None, description="情感标签: Positive/Neutral/Negative"),
    has_chat: Optional[str] = Query(None, description="聊天状态: yes/no"),
    use_default_filter: bool = Query(True, description="使用默认筛选(紧急/高优先级 OR 负面情感)")
):
    """
    导出优先关注客户为CSV文件(带AI画像)

    导出字段:
    - user_nick, buyer_type, follow_priority, sentiment_label, dominant_intent
    - rfm_segment, l6m_netsales, l1y_netsales, l1y_refund_rate, last_purchase_date
    - key_interests, pain_points, recommended_action (AI画像, 分号分隔)
    """
    try:
        from fastapi.responses import StreamingResponse
        from io import StringIO
        import csv
        from datetime import datetime

        # 获取所有客户(不分页)
        customers = analyzer.get_priority_customers(
            channel=channel,
            buyer_type=buyer_type,
            follow_priority=follow_priority,
            sentiment_label=sentiment_label,
            has_chat=has_chat,
            use_default_filter=use_default_filter,
            limit=1000,  # 最多导出1000条
            offset=0
        )

        # 创建CSV
        output = StringIO()
        writer = csv.writer(output)

        # CSV表头
        headers = [
            'user_nick', 'channel', 'buyer_type', 'follow_priority',
            'sentiment_label', 'dominant_intent', 'rfm_segment',
            'l6m_netsales', 'l1y_netsales', 'l1y_refund_rate',
            'last_purchase_date', 'has_chat',
            'key_interests', 'pain_points', 'recommended_action'
        ]
        writer.writerow(headers)

        # 写入数据
        for customer in customers:
            # 处理JSON数组字段 (转换为分号分隔的字符串)
            key_interests = customer.get('persona_key_interests', [])
            if isinstance(key_interests, list):
                key_interests = '; '.join(key_interests) if key_interests else 'N/A'
            elif key_interests is None:
                key_interests = 'N/A'

            pain_points = customer.get('persona_pain_points', [])
            if isinstance(pain_points, list):
                pain_points = '; '.join(pain_points) if pain_points else 'N/A'
            elif pain_points is None:
                pain_points = 'N/A'

            recommended_action = customer.get('persona_recommended_action') or 'N/A'

            row = [
                customer.get('buyer_nick', ''),
                customer.get('channel', ''),
                customer.get('buyer_type', ''),
                customer.get('follow_priority', ''),
                customer.get('sentiment_label', ''),
                customer.get('dominant_intent', ''),
                customer.get('rfm_segment', ''),
                customer.get('l6m_netsales', 0),
                customer.get('l1y_netsales', 0),
                f"{customer.get('l1y_refund_rate', 0) * 100:.1f}%",
                customer.get('last_purchase_date', ''),
                'Yes' if customer.get('has_chat') else 'No',
                key_interests,
                pain_points,
                recommended_action
            ]
            writer.writerow(row)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'priority_customers_{timestamp}.csv'

        # 返回CSV响应 - 使用utf-8-sig编码自动添加BOM以支持Excel正确显示中文
        output.seek(0)
        csv_content = output.getvalue()
        return StreamingResponse(
            iter([csv_content.encode('utf-8-sig')]),  # utf-8-sig会自动添加BOM (EF BB BF)
            media_type='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/{user_nick}/orders")
async def get_buyer_orders(
    user_nick: str,
    time_range: str = Query('1y', description="时间范围: 7d/15d/30d/90d/1y/all"),
    limit: int = Query(50, ge=1, le=1000)
) -> List[Dict[str, Any]]:
    """
    获取买家订单历史(使用订单明细预计算表 - 超快!)

    性能: < 0.5秒 (使用索引查询)

    支持的时间范围:
    - 7d: 近7天
    - 15d: 近15天
    - 30d: 近30天
    - 90d: 近90天
    - 1y: 近1年(默认)
    - all: 全部历史
    """
    try:
        from backend.database import Database
        from backend.config import settings
        import pymysql

        # 默认使用aliyunDB数据库
        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        # 构建时间范围条件
        time_conditions = {
            '7d': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 7 DAY)",
            '15d': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 15 DAY)",
            '30d': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 30 DAY)",
            '90d': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 90 DAY)",
            '1y': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 1 YEAR)",
            'all': ""
        }

        time_condition = time_conditions.get(time_range, time_conditions['1y'])

        # 使用订单明细预计算表查询
        query = f"""
            SELECT
                订单号, 子订单号, 商品名称, category as category,
                成交总金额, 退款金额, 退款类型,
                FP_MD, 图片地址, 最后付款时间, 件数,
                (成交总金额 - IFNULL(退款金额, 0)) as netsales
            FROM target_buyer_orders
            WHERE 买家昵称 = %s
              {time_condition}
            ORDER BY 最后付款时间 DESC
            LIMIT %s
        """

        params = [user_nick, limit]
        orders = db.execute_query(query, params)

        return orders
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/{user_nick}/chats")
async def get_buyer_chats(
    user_nick: str,
    limit: int = Query(100, ge=1, le=1000)
) -> List[Dict[str, Any]]:
    """
    获取买家聊天记录

    注意: 这个查询会比较多,建议限制返回数量
    """
    try:
        from backend.database import Database
        from backend.config import settings

        # 默认使用aliyunDB数据库
        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        query, params = BuyerQueries.get_chat_messages(user_nick, limit)
        chats = db.execute_query(query, params)

        return chats
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


async def _add_ai_analysis(profile: Dict[str, Any]) -> Dict[str, Any]:
    """
    添加AI生成的买家洞察（使用增强版AI分析系统）

    特性:
    - 两阶段推理: 证据提取 → 画像推理
    - 多级降级: DeepSeek-R1 → DeepSeek-Chat → Zhipu GLM-4 → 规则引擎
    - 成本监控: 实时记录API调用和成本
    - 智能缓存: 30天缓存，减少重复调用

    性能: 3-10秒 (取决于模型和数据复杂度)
    """
    try:
        import logging
        # 获取buyer_nick - 优先使用buyer_nick字段，兼容user_nick
        user_nick = profile.get("buyer_nick") or profile.get("user_nick")
        logging.info(f"[DEBUG AI] user_nick={user_nick}")

        # 1. 获取聊天记录(最近30条，增强分析需要更多数据)
        from backend.database import Database
        from backend.config import settings
        from backend.database import BuyerQueries

        # 默认使用aliyunDB数据库
        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        # 获取聊天记录
        query, params = BuyerQueries.get_chat_messages(user_nick, limit=30)
        chats = db.execute_query(query, params)
        logging.info(f"[DEBUG AI] fetched {len(chats)} chats")

        # 2. 获取订单记录（用于行为分析）
        # 注意: 使用原始表dunhill_t01_trade_line获取完整历史订单
        # target_buyer_orders表只有近3个月数据，会导致AI分析数据不完整
        orders_query = """
            SELECT
                订单号, 商品名称 as commodity_name, category,
                成交总金额 as payment, 退款金额, 退款类型 as refund_status,
                最后付款时间 as pay_time
            FROM dunhill_t01_trade_line
            WHERE 买家昵称 = %s
            ORDER BY 最后付款时间 DESC
            LIMIT 50
        """
        orders = db.execute_query(orders_query, [user_nick])
        logging.info(f"[DEBUG AI] fetched {len(orders)} orders")

        # 2.5 获取场外信息记录（用于补充上下文）
        from backend.analytics.external_analyzer import get_external_analyzer
        external_analyzer = get_external_analyzer()
        external_records = external_analyzer.get_records_by_user(user_nick, limit=20)
        logging.info(f"[DEBUG AI] fetched {len(external_records)} external records")

        # 3. 准备profile数据（传递所有预计算表字段，除了updated_at）
        # 确保AI能够基于完整的真实数据进行分析

        # ⚠️ 预先计算距今天数，避免AI自己计算时产生幻觉
        from datetime import datetime as dt
        today = dt.now()

        last_purchase_date = profile.get('last_purchase_date')
        first_purchase_date = profile.get('first_purchase_date')
        last_chat_date = profile.get('last_chat_date')

        days_since_last_purchase = 0
        days_since_last_chat = 0
        avg_repurchase_interval_days = 0

        if last_purchase_date:
            try:
                if isinstance(last_purchase_date, str):
                    last_purchase_date = dt.strptime(last_purchase_date[:19], '%Y-%m-%d %H:%M:%S')
                days_since_last_purchase = (today - last_purchase_date).days
            except:
                pass

        if last_chat_date:
            try:
                if isinstance(last_chat_date, str):
                    last_chat_date = dt.strptime(last_chat_date[:19], '%Y-%m-%d %H:%M:%S')
                days_since_last_chat = (today - last_chat_date).days
            except:
                pass

        # 计算平均复购间隔
        total_orders = int(profile.get('total_orders', 0))
        if first_purchase_date and last_purchase_date and total_orders > 1:
            try:
                if isinstance(first_purchase_date, str):
                    first_purchase_date = dt.strptime(first_purchase_date[:19], '%Y-%m-%d %H:%M:%S')
                if isinstance(last_purchase_date, str):
                    last_purchase_date = dt.strptime(last_purchase_date[:19], '%Y-%m-%d %H:%M:%S')
                days_span = (last_purchase_date - first_purchase_date).days
                avg_repurchase_interval_days = round(days_span / (total_orders - 1)) if days_span > 0 else 0
            except:
                pass

        profile_data = {
            # 基本信息（直接从profile传递所有字段）
            "user_nick": user_nick,
            "buyer_nick": profile.get('buyer_nick'),
            "channel": profile.get('channel'),
            "buyer_type": profile.get('buyer_type'),
            "is_smoker": profile.get('is_smoker', 0),
            "is_vic": profile.get('is_vic', 0),
            "vip_level": profile.get('vip_level', 'Non-VIP'),
            "client_monthly_tag": profile.get('client_monthly_tag'),

            # 历史消费指标
            "historical_gmv": float(profile.get('historical_gmv', 0)),
            "historical_refund": float(profile.get('historical_refund', 0)),
            "historical_net_sales": float(profile.get('historical_net_sales', 0)),
            "total_orders": int(profile.get('total_orders', 0)),
            "total_net_orders": int(profile.get('total_net_orders', 0)),
            "refund_rate": float(profile.get('refund_rate', 0)),

            # 时间维度指标
            "first_purchase_date": str(profile.get('first_purchase_date', '')) if profile.get('first_purchase_date') else '',
            "last_purchase_date": str(profile.get('last_purchase_date', '')) if profile.get('last_purchase_date') else '',

            # ⚠️ 预计算的天数字段（AI直接使用，不要自己计算！）
            "days_since_last_purchase": days_since_last_purchase,
            "days_since_last_chat": days_since_last_chat,
            "avg_repurchase_interval_days": avg_repurchase_interval_days,

            # 滚动24个月指标
            "rolling_24m_netsales": float(profile.get('rolling_24m_netsales', 0)),
            "rolling_24m_orders": int(profile.get('rolling_24m_orders', 0)),

            # L6M指标
            "l6m_gmv": float(profile.get('l6m_gmv', 0)),
            "l6m_netsales": float(profile.get('l6m_netsales', 0)),
            "l6m_orders": int(profile.get('l6m_orders', 0)),
            "l6m_refund_rate": float(profile.get('l6m_refund_rate', 0)),

            # L1Y指标
            "l1y_gmv": float(profile.get('l1y_gmv', 0)),
            "l1y_netsales": float(profile.get('l1y_netsales', 0)),
            "l1y_orders": int(profile.get('l1y_orders', 0)),
            "l1y_refund_rate": float(profile.get('l1y_refund_rate', 0)),

            # 折扣和敏感度
            "discount_ratio": float(profile.get('discount_ratio', 0)),
            "discount_sensitivity": profile.get('discount_sensitivity', '未知'),

            # 聊天行为指标
            "chat_frequency_days": int(profile.get('chat_frequency_days', 0)),
            "first_chat_date": str(profile.get('first_chat_date')) if profile.get('first_chat_date') else None,
            "last_chat_date": str(profile.get('last_chat_date')) if profile.get('last_chat_date') else None,
            "l30d_chat_frequency_days": int(profile.get('l30d_chat_frequency_days', 0)),
            "l3m_chat_frequency_days": int(profile.get('l3m_chat_frequency_days', 0)),
            "avg_chat_interval_days": float(profile.get('avg_chat_interval_days', 0)),

            # 风险和偏好
            "churn_risk": profile.get('churn_risk', '未知'),
            "city": profile.get('city', 'Unknown'),
            "top_category": profile.get('top_category', 'Unknown'),
            "second_category": profile.get('second_category'),
            "third_category": profile.get('third_category'),

            # AI分析专用字段
            "chat_history": chats,  # 传递聊天记录用于文本分析
            "external_records": external_records,  # 场外信息记录
            "total_refund_count": int(float(profile.get('historical_refund', 0)) / 1000) if float(profile.get('historical_refund', 0)) > 0 else 0  # 估算退款次数
        }

        # 4. 调用增强版AI分析器（自动选择最优模型和降级策略）
        import logging
        logging.info(f"[AI] 开始分析客户 {user_nick}, VIP等级: {profile_data['vip_level']}")

        # 检查是否应该使用缓存
        should_update = ai_orchestrator.should_update_persona(user_nick, profile_data)
        logging.info(f"[AI] should_update_persona={should_update}, ai_enable_cache={settings.ai_enable_cache}")

        ai_analysis = ai_orchestrator.analyze_buyer_persona(
            buyer_nick=user_nick,
            profile=profile_data,
            chats=chats,
            orders=orders
        )

        # 5. 添加AI分析结果到profile
        profile["ai_analysis"] = ai_analysis

        # 5.5 添加场外信息到profile（供前端展示）
        if external_records:
            profile["external_records"] = external_records

        # 记录使用的分析方法
        method = ai_analysis.get('analysis_method', 'Unknown')
        confidence = ai_analysis.get('confidence_level', 'Unknown')
        from_cache = ai_analysis.get('from_cache', False)
        logging.info(f"[AI] 分析完成 {user_nick}, 方法: {method}, 置信度: {confidence}, from_cache: {from_cache}")

        return profile

    except Exception as e:
        import logging
        logging.error(f"AI分析失败 for user {user_nick}: {e}", exc_info=True)
        # AI分析失败不影响基本信息返回
        profile["ai_analysis"] = {
            "summary": f"AI分析暂时不可用: {str(e)[:100]}",
            "key_interests": [],
            "pain_points": [],
            "recommended_action": "建议根据买家历史购买情况制定个性化跟进方案",
            "analysis_method": "Error",
            "error": str(e)
        }
        return profile


@router.post("/buyers/{user_nick}/force-refresh")
async def force_refresh_analysis(
    user_nick: str,
    refresh_type: str = "all",  # "persona" | "sentiment" | "all"
    reanalyze: bool = Query(True, description="是否立即重新分析")
) -> Dict[str, Any]:
    """
    强制刷新AI分析结果

    使用场景:
    - AI返回了错误结果
    - 分析逻辑升级后需要重新分析
    - 数据异常需要重新计算

    Args:
        user_nick: 买家昵称
        refresh_type: 刷新类型
            - "persona": 仅刷新客户画像
            - "sentiment": 仅刷新情感/意图分析
            - "all": 刷新全部
        reanalyze: 是否在清除缓存后立即重新分析（默认True）

    Returns:
        {
            "buyer_nick": str,
            "refresh_type": str,
            "cleared": list,
            "reanalyzed": list,
            "ai_provider": str,  # 如果重新分析了情感，返回使用的AI模型
            "message": str
        }
    """
    try:
        from backend.ai.analyzer_orchestrator import get_analyzer_orchestrator
        from backend.ai.batch_analyzer import get_batch_analyzer
        from backend.database import Database, BuyerQueries
        from backend.config import settings
        import logging

        orchestrator = get_analyzer_orchestrator()
        batch_analyzer = get_batch_analyzer()

        cleared = []
        reanalyzed = []
        ai_provider = None

        # 1. 清除缓存
        if refresh_type in ["persona", "all"]:
            orchestrator.force_refresh(user_nick)
            cleared.append("画像")

        if refresh_type in ["sentiment", "all"]:
            batch_analyzer.force_refresh(user_nick)
            cleared.append("情感")

        # 2. 如果需要重新分析
        if reanalyze:
            db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
            db = Database(db_name=db_name)

            # 获取聊天记录
            query, params = BuyerQueries.get_chat_messages(user_nick, limit=50)
            chats = db.execute_query(query, params)

            if refresh_type in ["sentiment", "all"]:
                # 重新进行情感/意图分析
                result = batch_analyzer.analyze_single_buyer(user_nick, chats)

                # 获取客户档案用于保存
                profile_query = "SELECT * FROM target_buyers_precomputed WHERE buyer_nick = %s"
                profile_result = db.execute_query(profile_query, [user_nick])
                profile = profile_result[0] if profile_result else None

                # 保存结果
                batch_analyzer.save_analysis_result(result, profile=profile)
                reanalyzed.append("情感")
                ai_provider = result.get('sentiment_method', 'unknown')
                logging.info(f"[ForceRefresh] 情感分析完成: {user_nick}, method={ai_provider}")

            if refresh_type in ["persona", "all"]:
                # 重新进行画像分析
                from backend.analytics.target_buyer_analyzer import TargetBuyerAnalyzer

                analyzer = TargetBuyerAnalyzer()
                profile = analyzer.get_buyer_profile(user_nick)

                if profile:
                    # 准备profile_data（包含预计算的天数字段）
                    from datetime import datetime as dt
                    today = dt.now()

                    last_purchase_date = profile.get('last_purchase_date')
                    first_purchase_date = profile.get('first_purchase_date')
                    last_chat_date = profile.get('last_chat_date')

                    days_since_last_purchase = 0
                    days_since_last_chat = 0
                    avg_repurchase_interval_days = 0

                    if last_purchase_date:
                        try:
                            if isinstance(last_purchase_date, str):
                                last_purchase_date = dt.strptime(last_purchase_date[:19], '%Y-%m-%d %H:%M:%S')
                            days_since_last_purchase = (today - last_purchase_date).days
                        except:
                            pass

                    if last_chat_date:
                        try:
                            if isinstance(last_chat_date, str):
                                last_chat_date = dt.strptime(last_chat_date[:19], '%Y-%m-%d %H:%M:%S')
                            days_since_last_chat = (today - last_chat_date).days
                        except:
                            pass

                    # 计算平均复购间隔
                    total_orders = int(profile.get('total_orders', 0))
                    if first_purchase_date and last_purchase_date and total_orders > 1:
                        try:
                            if isinstance(first_purchase_date, str):
                                first_purchase_date = dt.strptime(first_purchase_date[:19], '%Y-%m-%d %H:%M:%S')
                            if isinstance(last_purchase_date, str):
                                last_purchase_date = dt.strptime(last_purchase_date[:19], '%Y-%m-%d %H:%M:%S')
                            days_span = (last_purchase_date - first_purchase_date).days
                            avg_repurchase_interval_days = round(days_span / (total_orders - 1)) if days_span > 0 else 0
                        except:
                            pass

                    # 构建profile_data
                    profile_data = {
                        "user_nick": user_nick,
                        "buyer_nick": profile.get('buyer_nick'),
                        "vip_level": profile.get('vip_level', 'Non-VIP'),
                        "buyer_type": profile.get('buyer_type'),
                        "city": profile.get('city', 'Unknown'),
                        "first_purchase_date": str(profile.get('first_purchase_date', '')),
                        "last_purchase_date": str(profile.get('last_purchase_date', '')),
                        "total_orders": int(profile.get('total_orders', 0)),
                        "historical_net_sales": float(profile.get('historical_net_sales', 0)),
                        "days_since_last_purchase": days_since_last_purchase,
                        "days_since_last_chat": days_since_last_chat,
                        "avg_repurchase_interval_days": avg_repurchase_interval_days,
                        "top_category": profile.get('top_category', 'Unknown'),
                        "churn_risk": profile.get('churn_risk', '未知'),
                        "chat_history": chats,
                    }

                    # 调用orchestrator进行分析
                    ai_result = orchestrator.analyze_buyer_persona(
                        buyer_nick=user_nick,
                        profile=profile_data,
                        chats=chats,
                        orders=[]  # 订单数据已在profile中
                    )

                    # 保存结果到缓存
                    orchestrator.cache_manager.set_persona(user_nick, ai_result, profile_data)

                    reanalyzed.append("画像")
                    persona_method = ai_result.get('analysis_method', 'unknown')
                    logging.info(f"[ForceRefresh] 画像分析完成: {user_nick}, method={persona_method}")
                else:
                    logging.warning(f"[ForceRefresh] 未找到客户档案: {user_nick}")

        # 构建返回消息
        if reanalyzed:
            message = f"已清除 {', '.join(cleared)} 缓存并重新分析完成"
        else:
            message = f"已清除 {', '.join(cleared)} 缓存，下次请求将重新分析"

        return {
            "buyer_nick": user_nick,
            "refresh_type": refresh_type,
            "cleared": cleared,
            "reanalyzed": reanalyzed,
            "ai_provider": ai_provider,
            "message": message
        }

    except Exception as e:
        import logging
        logging.error(f"[ForceRefresh] 失败: {e}")
        raise HTTPException(status_code=500, detail=f"强制刷新失败: {str(e)}")


@router.post("/buyers/{user_nick}/analyze-async")
async def analyze_buyer_async(
    user_nick: str
) -> Dict[str, str]:
    """
    异步AI分析 - 返回task_id供轮询

    性能: < 0.1秒（立即返回，后台处理）

    Args:
        user_nick: 买家昵称

    Returns:
        {
            "task_id": str,
            "status": "pending",
            "message": str
        }
    """
    try:
        import logging
        from backend.ai.task_queue import get_task_queue
        from backend.database import Database, BuyerQueries
        from backend.config import settings

        # 1. 获取买家档案
        profile = analyzer.get_buyer_profile(user_nick)
        if not profile:
            raise HTTPException(status_code=404, detail=f"买家 {user_nick} 不存在")

        # 2. 获取聊天记录
        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)
        query, params = BuyerQueries.get_chat_messages(user_nick, limit=30)
        chats = db.execute_query(query, params)

        # 3. 获取订单记录
        orders_query = """
            SELECT
                订单号, 商品名称 as commodity_name, category,
                成交总金额 as payment, 退款金额, 退款类型 as refund_status,
                最后付款时间 as pay_time
            FROM target_buyer_orders
            WHERE 买家昵称 = %s
            ORDER BY 最后付款时间 DESC
            LIMIT 50
        """
        orders = db.execute_query(orders_query, [user_nick])

        # 4. 准备profile数据（与_add_ai_analysis相同）
        from datetime import datetime
        profile_data = {
            "user_nick": user_nick,
            "buyer_nick": profile.get('buyer_nick'),
            "channel": profile.get('channel'),
            "buyer_type": profile.get('buyer_type'),
            "is_smoker": profile.get('is_smoker', 0),
            "is_vic": profile.get('is_vic', 0),
            "vip_level": profile.get('vip_level', 'Non-VIP'),
            "client_monthly_tag": profile.get('client_monthly_tag'),
            "historical_gmv": float(profile.get('historical_gmv', 0)),
            "historical_refund": float(profile.get('historical_refund', 0)),
            "historical_net_sales": float(profile.get('historical_net_sales', 0)),
            "total_orders": int(profile.get('total_orders', 0)),
            "total_net_orders": int(profile.get('total_net_orders', 0)),
            "refund_rate": float(profile.get('refund_rate', 0)),
            "first_purchase_date": profile.get('first_purchase_date', ''),
            "last_purchase_date": profile.get('last_purchase_date', ''),
            "rolling_24m_netsales": float(profile.get('rolling_24m_netsales', 0)),
            "rolling_24m_orders": int(profile.get('rolling_24m_orders', 0)),
            "l6m_gmv": float(profile.get('l6m_gmv', 0)),
            "l6m_netsales": float(profile.get('l6m_netsales', 0)),
            "l6m_orders": int(profile.get('l6m_orders', 0)),
            "l6m_refund_rate": float(profile.get('l6m_refund_rate', 0)),
            "l1y_gmv": float(profile.get('l1y_gmv', 0)),
            "l1y_netsales": float(profile.get('l1y_netsales', 0)),
            "l1y_orders": int(profile.get('l1y_orders', 0)),
            "l1y_refund_rate": float(profile.get('l1y_refund_rate', 0)),
            "discount_ratio": float(profile.get('discount_ratio', 0)),
            "discount_sensitivity": profile.get('discount_sensitivity', '未知'),
            "chat_frequency_days": int(profile.get('chat_frequency_days', 0)),
            "first_chat_date": profile.get('first_chat_date'),
            "last_chat_date": profile.get('last_chat_date'),
            "l30d_chat_frequency_days": int(profile.get('l30d_chat_frequency_days', 0)),
            "l3m_chat_frequency_days": int(profile.get('l3m_chat_frequency_days', 0)),
            "avg_chat_interval_days": float(profile.get('avg_chat_interval_days', 0)),
            "churn_risk": profile.get('churn_risk', '未知'),
            "city": profile.get('city', 'Unknown'),
            "top_category": profile.get('top_category', 'Unknown'),
            "second_category": profile.get('second_category'),
            "third_category": profile.get('third_category'),
            "chat_history": chats,
            "total_refund_count": int(float(profile.get('historical_refund', 0)) / 1000) if float(profile.get('historical_refund', 0)) > 0 else 0
        }

        # 5. 加入任务队列
        task_queue = get_task_queue()
        task_id = await task_queue.enqueue(
            buyer_nick=user_nick,
            profile=profile_data,
            chats=chats,
            orders=orders,
            analyzer_orchestrator=ai_orchestrator
        )

        logging.info(f"[Async AI] 任务已创建: {task_id} for {user_nick}")

        return {
            "task_id": task_id,
            "status": "pending",
            "message": "AI分析任务已创建，请使用task_id查询结果"
        }

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"创建异步AI任务失败 for {user_nick}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tasks/{task_id}")
async def get_task_status(
    task_id: str
) -> Dict[str, Any]:
    """
    查询异步任务状态和结果

    Args:
        task_id: 任务ID

    Returns:
        {
            "task_id": str,
            "status": str,  # pending/processing/completed/failed
            "created_at": str,
            "started_at": str,
            "completed_at": str,
            "result": dict or None,
            "error": str or None,
            "buyer_nick": str
        }
    """
    try:
        from backend.ai.task_queue import get_task_queue

        task_queue = get_task_queue()
        task = task_queue.get_task(task_id)

        if not task:
            raise HTTPException(status_code=404, detail=f"任务 {task_id} 不存在")

        return task

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tasks/stats")
async def get_task_queue_stats() -> Dict[str, Any]:
    """
    获取任务队列统计

    Returns:
        {
            "total_tasks": int,
            "pending": int,
            "processing": int,
            "completed": int,
            "failed": int,
            "processing_count": int,
            "queue_size": int
        }
    """
    try:
        from backend.ai.task_queue import get_task_queue

        task_queue = get_task_queue()
        stats = task_queue.get_stats()

        return stats

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/{user_nick}/orders/export")
async def export_buyer_orders_excel(
    user_nick: str,
    time_range: str = Query('all', description="时间范围: 7d/15d/30d/90d/1y/all")
):
    """
    导出买家订单历史为 Excel 文件

    Args:
        user_nick: 买家昵称
        time_range: 时间范围 (默认 all 导出全部)

    Returns:
        Excel 文件下载
    """
    try:
        from backend.database import Database
        from backend.config import settings
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # 获取数据
        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        time_conditions = {
            '7d': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 7 DAY)",
            '15d': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 15 DAY)",
            '30d': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 30 DAY)",
            '90d': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 90 DAY)",
            '1y': "AND 最后付款时间 >= DATE_SUB(NOW(), INTERVAL 1 YEAR)",
            'all': ""
        }

        time_condition = time_conditions.get(time_range, time_conditions['all'])

        query = f"""
            SELECT
                订单号, 子订单号, 商品名称, category as 品类,
                成交总金额, 退款金额, 退款类型,
                FP_MD, 最后付款时间, 件数,
                (成交总金额 - IFNULL(退款金额, 0)) as 净销售额
            FROM target_buyer_orders
            WHERE 买家昵称 = %s
              {time_condition}
            ORDER BY 最后付款时间 DESC
        """

        orders = db.execute_query(query, [user_nick])

        if not orders:
            raise HTTPException(status_code=404, detail=f"买家 {user_nick} 暂无订单记录")

        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        # Excel sheet名称不能包含: \ / ? * [ ]
        safe_sheet_name = ''.join(c if c not in '\\/?*[]' else '_' for c in user_nick[:20])
        ws.title = f"订单历史_{safe_sheet_name}"

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = ["订单号", "子订单号", "商品名称", "品类", "成交金额", "退款金额",
                   "退款类型", "FP_MD", "付款时间", "件数", "净销售额"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, order in enumerate(orders, 2):
            ws.cell(row=row_idx, column=1, value=order.get("订单号", ""))
            ws.cell(row=row_idx, column=2, value=order.get("子订单号", ""))
            ws.cell(row=row_idx, column=3, value=order.get("商品名称", ""))
            ws.cell(row=row_idx, column=4, value=order.get("品类", ""))
            ws.cell(row=row_idx, column=5, value=float(order.get("成交总金额", 0) or 0))
            ws.cell(row=row_idx, column=6, value=float(order.get("退款金额", 0) or 0))
            ws.cell(row=row_idx, column=7, value=order.get("退款类型", ""))
            ws.cell(row=row_idx, column=8, value=order.get("FP_MD", ""))
            ws.cell(row=row_idx, column=9, value=str(order.get("最后付款时间", "")))
            ws.cell(row=row_idx, column=10, value=int(order.get("件数", 1) or 1))
            ws.cell(row=row_idx, column=11, value=float(order.get("净销售额", 0) or 0))

            # 添加边框
            for col in range(1, 12):
                ws.cell(row=row_idx, column=col).border = thin_border

        # 设置列宽
        column_widths = [20, 20, 40, 15, 12, 12, 15, 10, 20, 8, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 汇总行
        summary_row = len(orders) + 2
        ws.cell(row=summary_row, column=1, value="合计")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)

        total_gmv = sum(float(o.get("成交总金额", 0) or 0) for o in orders)
        total_refund = sum(float(o.get("退款金额", 0) or 0) for o in orders)
        total_netsales = sum(float(o.get("净销售额", 0) or 0) for o in orders)

        ws.cell(row=summary_row, column=5, value=total_gmv)
        ws.cell(row=summary_row, column=6, value=total_refund)
        ws.cell(row=summary_row, column=11, value=total_netsales)

        for col in range(1, 12):
            ws.cell(row=summary_row, column=col).font = Font(bold=True)
            ws.cell(row=summary_row, column=col).border = thin_border

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名（使用URL编码处理中文）
        from datetime import datetime
        from urllib.parse import quote
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 文件名格式: orders-{user_nick}{timestamp}.xlsx
        filename = f"orders-{user_nick}{timestamp}.xlsx"
        encoded_filename = quote(filename)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"导出Excel失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# RFM Customer Classification Endpoints
# ============================================

@router.get("/buyers/rfm/summary")
async def get_rfm_summary() -> Dict[str, Any]:
    """
    获取RFM客户分类汇总统计

    Returns:
        {
            "segments": [
                {"segment": str, "count": int, "avg_netsales": float, "avg_orders": float}
            ],
            "total_buyers": int
        }
    """
    try:
        from backend.database import Database
        from backend.config import settings

        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        query = """
            SELECT
                rfm_segment,
                COUNT(*) as count,
                ROUND(AVG(historical_net_sales), 2) as avg_netsales,
                ROUND(AVG(total_orders), 2) as avg_orders,
                ROUND(AVG(rfm_recency_score), 2) as avg_r,
                ROUND(AVG(rfm_frequency_score), 2) as avg_f,
                ROUND(AVG(rfm_monetary_score), 2) as avg_m
            FROM target_buyers_precomputed
            WHERE rfm_segment IS NOT NULL
            GROUP BY rfm_segment
            ORDER BY
                CASE rfm_segment
                    WHEN '重要价值客户' THEN 1
                    WHEN '重要发展客户' THEN 2
                    WHEN '重要保持客户' THEN 3
                    WHEN '一般价值客户' THEN 4
                    WHEN '潜力客户' THEN 5
                    WHEN '流失预警' THEN 6
                    WHEN '已流失' THEN 7
                    ELSE 8
                END
        """

        segments = db.execute_query(query)

        total_query = "SELECT COUNT(*) as total FROM target_buyers_precomputed WHERE rfm_segment IS NOT NULL"
        total_result = db.execute_query(total_query)
        total_buyers = total_result[0].get('total', 0) if total_result else 0

        return {
            "segments": segments,
            "total_buyers": total_buyers
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/rfm/{rfm_segment}")
async def get_buyers_by_rfm_segment(
    rfm_segment: str,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    按RFM客户分类筛选买家

    性能: < 0.1秒 (使用rfm_segment索引)

    Args:
        rfm_segment: RFM客户分类
            - 重要价值客户: 高频高消费且最近活跃
            - 重要发展客户: 高消费但频次低
            - 重要保持客户: 曾经高频高消费但近期沉默
            - 一般价值客户: 中等活跃度
            - 潜力客户: 新客户，有潜力
            - 流失预警: 曾经有价值但可能流失
            - 已流失: 长期无活动
    """
    try:
        valid_segments = [
            '重要价值客户', '重要发展客户', '重要保持客户',
            '一般客户', '潜力客户', '流失预警', '已流失'
        ]

        if rfm_segment not in valid_segments:
            raise HTTPException(
                status_code=400,
                detail=f"无效的rfm_segment: {rfm_segment}. 必须是: {', '.join(valid_segments)}"
            )

        buyers = analyzer.get_buyers_by_rfm_segment(rfm_segment, limit, offset)
        return buyers

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# Sentiment & Intent Analysis Endpoints
# ============================================

@router.post("/ai/batch-analyze")
async def start_batch_analysis(
    buyer_limit: int = Query(100, ge=1, le=500, description="Maximum buyers to analyze")
) -> Dict[str, Any]:
    """
    启动批量AI情绪/意图分析任务

    增量更新模式：只分析有新增聊天记录的客户

    Args:
        buyer_limit: 最大分析客户数 (1-500)

    Returns:
        {
            "task_id": str,
            "status": "pending",
            "message": str
        }
    """
    try:
        from backend.ai.batch_analyzer import get_batch_analyzer

        batch_analyzer = get_batch_analyzer()
        task_id = batch_analyzer.start_batch_analysis(buyer_limit=buyer_limit)

        return {
            "task_id": task_id,
            "status": "pending",
            "message": f"批量分析任务已创建，将分析最多 {buyer_limit} 个客户"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ai/batch-status/{task_id}")
async def get_batch_analysis_status(
    task_id: str
) -> Dict[str, Any]:
    """
    查询批量分析任务状态

    Args:
        task_id: 任务ID

    Returns:
        {
            "task_id": str,
            "status": str (pending/running/completed/failed),
            "total_buyers": int,
            "processed_buyers": int,
            "failed_buyers": int,
            "progress_percent": float,
            "started_at": str,
            "completed_at": str,
            "error_message": str
        }
    """
    try:
        from backend.ai.batch_analyzer import get_batch_analyzer

        batch_analyzer = get_batch_analyzer()
        status = batch_analyzer.get_task_status(task_id)

        if not status:
            raise HTTPException(status_code=404, detail=f"任务 {task_id} 不存在")

        return status

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/sentiment/{label}")
async def get_buyers_by_sentiment(
    label: str,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    按情绪标签筛选客户

    性能: < 0.1秒 (使用sentiment_label索引)
    更新: 2026-03-19 优先使用缓存表实时数据

    Args:
        label: 情绪标签 (Positive/Neutral/Negative)
    """
    try:
        label_upper = label.capitalize()
        if label_upper not in ['Positive', 'Neutral', 'Negative']:
            raise HTTPException(
                status_code=400,
                detail=f"无效的sentiment_label: {label}. 必须是 Positive, Neutral 或 Negative"
            )

        from backend.database import Database
        from backend.config import settings

        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        query = """
            SELECT
                tb.buyer_nick, tb.channel, tb.buyer_type, tb.vip_level,
                tb.historical_net_sales, tb.total_orders, tb.last_purchase_date,
                COALESCE(cache.sentiment_label, tb.sentiment_label) AS sentiment_label,
                COALESCE(cache.sentiment_score, tb.sentiment_score) AS sentiment_score,
                COALESCE(cache.dominant_intent, tb.dominant_intent) AS dominant_intent,
                tb.rfm_segment, tb.churn_risk
            FROM target_buyers_precomputed tb
            LEFT JOIN buyer_ai_analysis_cache cache ON tb.buyer_nick = cache.buyer_nick
            WHERE COALESCE(cache.sentiment_label, tb.sentiment_label) = %s
            ORDER BY tb.historical_net_sales DESC
            LIMIT %s OFFSET %s
        """

        buyers = db.execute_query(query, [label_upper, limit, offset])
        return buyers

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/intent/{intent}")
async def get_buyers_by_intent(
    intent: str,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    按主要意图筛选客户

    性能: < 0.1秒 (使用dominant_intent索引)
    更新: 2026-03-19 优先使用缓存表实时数据

    Args:
        intent: 主要意图
            - Pre-sale Inquiry: 售前咨询
            - Post-sale Support: 售后支持
            - Logistics: 物流查询
            - Usage Guide: 使用指南
            - Complaint: 投诉
    """
    try:
        valid_intents = [
            'Pre-sale Inquiry', 'Post-sale Support',
            'Logistics', 'Usage Guide', 'Complaint', 'Unknown'
        ]

        if intent not in valid_intents:
            raise HTTPException(
                status_code=400,
                detail=f"无效的intent: {intent}. 必须是: {', '.join(valid_intents)}"
            )

        from backend.database import Database
        from backend.config import settings

        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        query = """
            SELECT
                tb.buyer_nick, tb.channel, tb.buyer_type, tb.vip_level,
                tb.historical_net_sales, tb.total_orders, tb.last_purchase_date,
                COALESCE(cache.sentiment_label, tb.sentiment_label) AS sentiment_label,
                COALESCE(cache.sentiment_score, tb.sentiment_score) AS sentiment_score,
                COALESCE(cache.dominant_intent, tb.dominant_intent) AS dominant_intent,
                tb.rfm_segment, tb.churn_risk
            FROM target_buyers_precomputed tb
            LEFT JOIN buyer_ai_analysis_cache cache ON tb.buyer_nick = cache.buyer_nick
            WHERE COALESCE(cache.dominant_intent, tb.dominant_intent) = %s
            ORDER BY tb.historical_net_sales DESC
            LIMIT %s OFFSET %s
        """

        buyers = db.execute_query(query, [intent, limit, offset])
        return buyers

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/buyers/follow-priority/{priority}")
async def get_buyers_by_follow_priority(
    priority: str,
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
) -> List[Dict[str, Any]]:
    """
    按跟进优先级筛选客户

    性能: < 0.1秒 (使用follow_priority索引)
    更新: 2026-03-19 优先使用缓存表实时数据

    Args:
        priority: 跟进优先级 (紧急/高/中/低)
    """
    try:
        if priority not in ['紧急', '高', '中', '低']:
            raise HTTPException(
                status_code=400,
                detail=f"无效的priority: {priority}. 必须是: 紧急, 高, 中 或 低"
            )

        from backend.database import Database
        from backend.config import settings

        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        query = """
            SELECT
                tb.buyer_nick, tb.channel, tb.buyer_type, tb.vip_level,
                tb.historical_net_sales, tb.total_orders, tb.last_purchase_date, tb.last_chat_date,
                COALESCE(cache.sentiment_label, tb.sentiment_label) AS sentiment_label,
                COALESCE(cache.sentiment_score, tb.sentiment_score) AS sentiment_score,
                COALESCE(cache.dominant_intent, tb.dominant_intent) AS dominant_intent,
                tb.rfm_segment, tb.churn_risk, tb.follow_priority
            FROM target_buyers_precomputed tb
            LEFT JOIN buyer_ai_analysis_cache cache ON tb.buyer_nick = cache.buyer_nick
            WHERE tb.follow_priority = %s
            ORDER BY
                CASE tb.vip_level
                    WHEN 'V3' THEN 1
                    WHEN 'V2' THEN 2
                    WHEN 'V1' THEN 3
                    WHEN 'V0' THEN 4
                    ELSE 5
                END,
                tb.historical_net_sales DESC
            LIMIT %s OFFSET %s
        """

        buyers = db.execute_query(query, [priority, limit, offset])
        return buyers

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/analytics/sentiment-summary")
async def get_sentiment_summary() -> Dict[str, Any]:
    """
    获取情绪分析汇总统计

    Returns:
        {
            "total_analyzed": int,
            "positive": {"count": int, "avg_score": float},
            "neutral": {"count": int, "avg_score": float},
            "negative": {"count": int, "avg_score": float},
            "distribution_percent": {...}
        }
    """
    try:
        from backend.ai.batch_analyzer import get_batch_analyzer

        batch_analyzer = get_batch_analyzer()
        summary = batch_analyzer.get_sentiment_summary()

        # Calculate percentages
        total = summary["total_analyzed"]
        if total > 0:
            summary["distribution_percent"] = {
                "positive": round(summary["positive"]["count"] / total * 100, 1),
                "neutral": round(summary["neutral"]["count"] / total * 100, 1),
                "negative": round(summary["negative"]["count"] / total * 100, 1)
            }

        return summary

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/analytics/intent-summary")
async def get_intent_summary() -> Dict[str, Any]:
    """
    获取意图分析汇总统计

    Returns:
        {
            "total_analyzed": int,
            "intents": {
                "Pre-sale Inquiry": int,
                "Post-sale Support": int,
                "Logistics": int,
                "Usage Guide": int,
                "Complaint": int
            }
        }
    """
    try:
        from backend.ai.batch_analyzer import get_batch_analyzer

        batch_analyzer = get_batch_analyzer()
        summary = batch_analyzer.get_intent_summary()

        return summary

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/analytics/follow-up-list")
async def get_follow_up_list(
    limit: int = Query(50, ge=1, le=200)
) -> Dict[str, List[Dict[str, Any]]]:
    """
    获取待跟进客户列表（按优先级分组）

    Returns:
        {
            "urgent": [...],  // 紧急跟进
            "high": [...],    // 高优先级
            "medium": [...]   // 中优先级
        }
    """
    try:
        from backend.database import Database
        from backend.config import settings

        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        result = {
            "urgent": [],
            "high": [],
            "medium": []
        }

        for priority in ['紧急', '高', '中']:
            query = """
                SELECT
                    tb.buyer_nick, tb.vip_level, tb.rfm_segment,
                    tb.historical_net_sales, tb.total_orders,
                    tb.last_purchase_date, tb.last_chat_date,
                    COALESCE(cache.sentiment_label, tb.sentiment_label) AS sentiment_label,
                    COALESCE(cache.dominant_intent, tb.dominant_intent) AS dominant_intent,
                    tb.churn_risk, tb.follow_priority
                FROM target_buyers_precomputed tb
                LEFT JOIN buyer_ai_analysis_cache cache ON tb.buyer_nick = cache.buyer_nick
                WHERE tb.follow_priority = %s
                ORDER BY
                    CASE tb.vip_level
                        WHEN 'V3' THEN 1
                        WHEN 'V2' THEN 2
                        WHEN 'V1' THEN 3
                        ELSE 4
                    END,
                    tb.historical_net_sales DESC
                LIMIT %s
            """

            buyers = db.execute_query(query, [priority, limit])

            # Convert to dict key format
            key_map = {'紧急': 'urgent', '高': 'high', '中': 'medium'}
            result[key_map[priority]] = buyers

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# 关键词分析 API
# ============================================================

@router.get("/keyword-analysis")
async def get_keyword_analysis(
    buyer_types: str = Query("ALL", description="客户类型，逗号分隔，如 'SMOKER,BOTH'，或 'ALL'"),
    category: str = Query(None, description="分类筛选，可选"),
    limit: int = Query(20, description="返回关键词数量限制")
) -> Dict[str, Any]:
    """
    获取关键词分析数据

    Args:
        buyer_types: 客户类型，支持多选（逗号分隔）或 'ALL'
        category: 分类筛选（可选）
        limit: 返回关键词数量限制

    Returns:
        {
            "category_distribution": [
                { "name": "赠品", "value": 112, "percentage": 10.1 },
                ...
            ],
            "keywords": [
                { "text": "赠品", "value": 45, "percentage": 23.1, "category": "赠品" },
                ...
            ],
            "total_messages": 416,
            "buyer_types": ["SMOKER", "BOTH"],
            "selected_category": null
        }
    """
    try:
        from backend.database import Database
        from backend.config import settings
        from collections import defaultdict

        db_name = settings.db_name_to_use if settings.db_name_to_use else 'aliyunDB'
        db = Database(db_name=db_name)

        # 解析客户类型
        buyer_type_list = [bt.strip().upper() for bt in buyer_types.split(',')]
        if 'ALL' in buyer_type_list:
            buyer_type_list = ['ALL']

        # 获取元数据（总消息数）
        total_messages = 0
        for bt in buyer_type_list:
            meta_query = """
                SELECT total_messages FROM keyword_analysis_meta WHERE buyer_type = %s
            """
            meta_result = db.execute_query(meta_query, [bt])
            if meta_result:
                total_messages += meta_result[0]['total_messages']

        # 获取分类分布
        category_distribution = []
        category_counts = defaultdict(int)

        for bt in buyer_type_list:
            cat_query = """
                SELECT category, SUM(count) as count, SUM(percentage) as percentage
                FROM category_distribution_cache
                WHERE buyer_type = %s
                GROUP BY category
                ORDER BY count DESC
            """
            cat_results = db.execute_query(cat_query, [bt])
            for row in cat_results:
                # 转换 Decimal 为 int
                category_counts[row['category']] += int(row['count'])

        # 计算百分比
        total_category_count = sum(category_counts.values())
        for cat_name, count in sorted(category_counts.items(), key=lambda x: -x[1]):
            percentage = round(count / total_category_count * 100, 1) if total_category_count > 0 else 0
            category_distribution.append({
                "name": cat_name,
                "value": count,
                "percentage": percentage
            })

        # 获取关键词
        keyword_query_parts = []
        query_params = []

        for bt in buyer_type_list:
            if category:
                keyword_query_parts.append("""
                    SELECT keyword, category, SUM(count) as count, SUM(percentage) as percentage
                    FROM keyword_analysis_cache
                    WHERE buyer_type = %s AND category = %s
                    GROUP BY keyword, category
                """)
                query_params.extend([bt, category])
            else:
                keyword_query_parts.append("""
                    SELECT keyword, category, SUM(count) as count, SUM(percentage) as percentage
                    FROM keyword_analysis_cache
                    WHERE buyer_type = %s
                    GROUP BY keyword, category
                """)
                query_params.append(bt)

        # 合并查询
        union_query = " UNION ALL ".join(keyword_query_parts)
        final_query = f"""
            SELECT keyword, category, SUM(count) as count, SUM(percentage) as percentage
            FROM ({union_query}) as combined
            GROUP BY keyword, category
            ORDER BY count DESC
            LIMIT %s
        """
        query_params.append(limit)

        keyword_results = db.execute_query(final_query, query_params)

        # 计算关键词总计数（用于重新计算百分比）
        total_keyword_count = sum(int(row['count']) for row in keyword_results) if keyword_results else 0

        keywords = []
        for row in keyword_results:
            # 转换 Decimal 为 int
            count = int(row['count'])
            # 重新计算百分比（基于合并后的数据）
            percentage = round(count / total_keyword_count * 100, 1) if total_keyword_count > 0 else 0
            keywords.append({
                "text": row['keyword'],
                "value": count,
                "percentage": percentage,
                "category": row['category']
            })

        return {
            "category_distribution": category_distribution,
            "keywords": keywords,
            "total_messages": total_messages,
            "buyer_types": buyer_type_list,
            "selected_category": category
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
